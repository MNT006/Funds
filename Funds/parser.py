"""
Parser for Thracian-style mutual fund NAV reports (.xlsx).

The REPORT sheet is a hierarchy:
  - section header rows in column 9 (e.g. "Акции във валута", "Облигации", …)
  - line-item rows below them with columns:
      col 10: issuer / security name
      col 11: ISIN
      col 12: ticker
      col 14: quantity
      col 17: dirty price (in the security's own currency)
      col 20: currency
      col 21: market value (in fund base currency, BGN)
      col 22: weight (as fraction 0..1)

Below the holdings block, the same column 9 holds NAV totals, cash sub-types,
and the per-unit pricing block.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl


# Maps the Bulgarian section headers to English asset-class labels used in the UI.
SECTION_TO_CLASS = {
    "дялове на кис": "Funds / ETFs",
    "инструменти на паричния пазар": "Money Market",
    "акции": "Equities",
    "деривативни финансови инструменти": "Derivatives",
    "опции": "Options",
    "права": "Rights",
    "облигации": "Bonds",
    "държавни ценни книжа": "Government Bonds",
}

# Bulgarian → English fund-name display label.
FUND_NAME_OVERRIDES = {
    "трейшън-алтернативен доход": "Thracian — Alternative Income",
    "трейшън-алтернативен  доход": "Thracian — Alternative Income",
    "трейшън-иновации и технологии": "Thracian — Innovation & Technology",
}


@dataclass
class Holding:
    asset_class: str
    security_name: str
    isin: str | None
    ticker: str | None
    quantity: float
    price_dirty: float | None
    currency: str | None
    market_value: float            # in fund base currency
    weight: float                  # fraction 0..1
    country: str | None = None     # filled by enrichment
    sector: str | None = None      # filled by enrichment


@dataclass
class CashLine:
    label: str
    market_value: float
    weight: float


@dataclass
class Liability:
    label: str
    market_value: float


@dataclass
class Fund:
    source_file: str
    fund_name: str                 # display label (English)
    fund_name_native: str          # original Bulgarian name
    reporting_date: date
    base_currency: str = "EUR"
    nav_total: float = 0.0
    units_outstanding: float = 0.0
    nav_per_unit: float = 0.0
    issue_price: float | None = None
    redemption_price_short: float | None = None
    redemption_price_long: float | None = None
    fx_rates: dict[str, float] = field(default_factory=dict)
    holdings: list[Holding] = field(default_factory=list)
    cash: list[CashLine] = field(default_factory=list)
    liabilities: list[Liability] = field(default_factory=list)
    receivables_total: float = 0.0
    deferred_expenses: float = 0.0
    total_liabilities: float = 0.0
    total_assets: float = 0.0


def _strip(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_na(v) -> bool:
    if v is None:
        return True
    s = _strip(v)
    return s == "" or s.upper() == "#N/A" or s == "0"


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.upper() == "#N/A":
        return None
    try:
        return float(s.replace(",", "").replace(" ", ""))
    except ValueError:
        return None


def _normalize_section(text: str) -> str | None:
    """Map a Bulgarian section header to an English asset-class label."""
    if not text:
        return None
    t = text.lower().strip().lstrip("0123456789. ")
    # Strip trailing currency qualifiers ("във валута", "в лева", …)
    t = re.sub(r"\s+(във валута|в лева|в евро)\s*$", "", t)
    for key, label in SECTION_TO_CLASS.items():
        if key in t:
            return label
    return None


def _country_from_isin(isin: str | None) -> str | None:
    # Real ISINs are exactly 12 chars (CC + 9 alnum + 1 check digit).
    # Anything else is likely a placeholder or option-contract code.
    if not isin or len(isin) != 12:
        return None
    prefix = isin[:2].upper()
    if not prefix.isalpha():
        return None
    return {
        "US": "United States",
        "BG": "Bulgaria",
        "GB": "United Kingdom",
        "AU": "Australia",
        "MH": "Marshall Islands",
        "KY": "Cayman Islands",
        "BM": "Bermuda",
        "DE": "Germany",
        "FR": "France",
        "NL": "Netherlands",
        "IE": "Ireland",
        "LU": "Luxembourg",
        "CA": "Canada",
        "JP": "Japan",
        "CH": "Switzerland",
    }.get(prefix, prefix)


def _english_fund_name(native: str) -> str:
    key = native.lower().strip()
    for needle, eng in FUND_NAME_OVERRIDES.items():
        if needle in key:
            return eng
    return native  # fall back to native if unknown


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Locate the holdings-table header row (the one containing 'Актив' /
    'ISIN код'). Returns (row_index, {col_name: col_index})."""
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [_strip(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if "isin код" in row_vals and "номинал / брой" in [v.replace("  ", " ") for v in row_vals]:
            cols = {}
            for c in range(1, ws.max_column + 1):
                v = _strip(ws.cell(r, c).value).lower().replace("  ", " ")
                if v:
                    cols[v] = c
            return r, cols
    raise ValueError("Could not locate holdings header row")


def _resolve(cols: dict[str, int], *aliases: str) -> int | None:
    for a in aliases:
        if a in cols:
            return cols[a]
    return None


def _parse_report_sheet(ws, source_file: str) -> Fund:
    header_row, cols = _find_header_row(ws)

    col_section = _resolve(cols, "актив")
    col_issuer = _resolve(cols, "емитент")
    col_isin = _resolve(cols, "isin код")
    col_ticker = _resolve(cols, "борсов код")
    col_qty = _resolve(cols, "номинал / брой")
    col_price = _resolve(cols, "мръсна цена в основна валута", "чиста цена в основна валута")
    col_ccy = _resolve(cols, "валута")
    col_mv = _resolve(cols, "стойност на актив")
    col_wt = _resolve(cols, "% актив")

    if not all([col_section, col_issuer, col_isin, col_ticker, col_qty, col_mv, col_wt]):
        raise ValueError(f"Missing expected columns; found {list(cols)}")

    # --- Fund name (row 1, anywhere)
    fund_name_native = ""
    for c in range(1, ws.max_column + 1):
        v = _strip(ws.cell(1, c).value)
        if "трейшън" in v.lower() or "thracian" in v.lower():
            fund_name_native = v
            break
    if not fund_name_native:
        # fallback to filename
        fund_name_native = Path(source_file).stem

    # --- Reporting date (row 2..6 col_mv area)
    reporting_date = None
    for r in range(2, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                reporting_date = v.date()
                break
        if reporting_date:
            break
    if reporting_date is None:
        # Try to derive from filename "04.05.2026_..."
        m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", Path(source_file).name)
        if m:
            d, mo, y = m.groups()
            reporting_date = date(int(y), int(mo), int(d))
        else:
            reporting_date = date.today()

    # --- Walk rows after header.
    fund = Fund(
        source_file=str(source_file),
        fund_name=_english_fund_name(fund_name_native),
        fund_name_native=fund_name_native,
        reporting_date=reporting_date,
    )

    current_class: str | None = None
    in_cash = False
    in_receivables = False
    in_liabilities = False

    for r in range(header_row + 1, ws.max_row + 1):
        section_text = _strip(ws.cell(r, col_section).value)
        # Many label rows put their text in col 10 (issuer column) instead
        label_col10 = _strip(ws.cell(r, col_issuer).value)
        mv = _to_float(ws.cell(r, col_mv).value)
        wt = _to_float(ws.cell(r, col_wt).value)
        if wt is not None and abs(wt) > 1.5:  # accidentally captured a percent number
            wt = wt / 100

        # --- Section transitions
        if section_text:
            low = section_text.lower()
            if "ценни книжа" in low and section_text.strip().startswith("1"):
                in_cash = in_receivables = in_liabilities = False
                current_class = None
                continue
            if "парични средства" in low and section_text.strip().startswith("2"):
                in_cash = True
                in_receivables = in_liabilities = False
                current_class = None
                continue
            if "нефинансови активи" in low or "вземания" in low:
                in_receivables = True
                in_cash = in_liabilities = False
                continue
            if "разходи за бъдещи периоди" in low:
                if mv is not None:
                    fund.deferred_expenses = mv
                current_class = None
                continue
            if "общо активи" in low:
                if mv is not None:
                    fund.total_assets = mv
                current_class = None
                continue
            # Order matters: "общо текущи пасиви" contains "текущи пасиви",
            # so check the longer phrase first.
            if "общо текущи пасиви" in low:
                if mv is not None:
                    fund.total_liabilities = mv
                in_liabilities = False
                continue
            if "текущи пасиви" in low:
                in_liabilities = True
                in_cash = in_receivables = False
                continue
            if "нетна стойност на активите" in low:
                if mv is not None:
                    fund.nav_total = mv
                continue

            # Otherwise, try as an asset-class section header.
            cls = _normalize_section(section_text)
            if cls:
                current_class = cls
                continue

        # --- Cash lines (label in col 10, value in col 21)
        if in_cash and label_col10 and mv is not None:
            fund.cash.append(CashLine(label=_clean_cash_label(label_col10),
                                      market_value=mv,
                                      weight=wt or 0.0))
            continue

        # --- Receivables (just sum to receivables_total)
        if in_receivables and label_col10 and mv is not None:
            if "вземания" in label_col10.lower() and not label_col10.strip().startswith("4."):
                # the "4.1 Вземания" parent is in col 9, items below are 4.1.x in col 10
                fund.receivables_total += 0  # parent sum captured separately
            else:
                fund.receivables_total += mv
            continue

        # --- Liabilities
        if in_liabilities and label_col10 and mv is not None:
            fund.liabilities.append(Liability(label=label_col10, market_value=mv))
            continue

        # --- Holding line (only meaningful while current_class is set)
        if current_class is None:
            continue

        issuer = _strip(ws.cell(r, col_issuer).value)
        isin = _strip(ws.cell(r, col_isin).value)
        ticker = _strip(ws.cell(r, col_ticker).value)
        qty = _to_float(ws.cell(r, col_qty).value) or 0
        price = _to_float(ws.cell(r, col_price).value) if col_price else None
        ccy = _strip(ws.cell(r, col_ccy).value) if col_ccy else None
        mv_val = _to_float(ws.cell(r, col_mv).value) or 0
        wt_val = wt or 0.0

        # Skip placeholder/empty rows.
        if (_is_na(issuer) and _is_na(isin) and _is_na(ticker)) or qty == 0 or mv_val == 0:
            continue

        fund.holdings.append(
            Holding(
                asset_class=current_class,
                security_name=issuer or ticker or (isin or "Unknown"),
                isin=isin if not _is_na(isin) else None,
                ticker=ticker if not _is_na(ticker) else None,
                quantity=qty,
                price_dirty=price,
                currency=(ccy if isinstance(ccy, str) and ccy and ccy != "0" else None),
                market_value=mv_val,
                weight=wt_val,
                country=_country_from_isin(isin if not _is_na(isin) else None),
            )
        )

    # --- After-table: per-unit + FX block
    _parse_footer(ws, fund, header_row)

    return fund


def _clean_cash_label(s: str) -> str:
    """Strip leading numbering like '3.2 ' and 'Парични средства по ' →
    keep the rest in original Bulgarian (it's short enough)."""
    s = re.sub(r"^\d+\.\d+\s*", "", s)
    return s


def _parse_footer(ws, fund: Fund, header_row: int) -> None:
    """Walk past the holdings table looking for FX, units, NAV/unit, prices."""
    for r in range(header_row, ws.max_row + 1):
        # Look at columns 9, 10, 11, 12 — the per-unit table is at col 10/11.
        c9 = _strip(ws.cell(r, 9).value).lower()
        c10 = _strip(ws.cell(r, 10).value).lower()
        c11_raw = ws.cell(r, 11).value
        c12_raw = ws.cell(r, 12).value

        if "валутни курсове" in c9:
            continue
        if c9 == "usd" or _strip(ws.cell(r, 9).value).upper() == "USD":
            rate = _to_float(c11_raw) or _to_float(c12_raw)
            if rate:
                fund.fx_rates["USD/EUR"] = rate
                # Convert to BGN/USD via the BGN/EUR fixed peg 1.95583
                fund.fx_rates["USD/BGN"] = rate * 1.95583
            continue

        if "общ брой дялове" in c10:
            v = _to_float(c11_raw)
            if v:
                fund.units_outstanding = v
        elif "нетна стойност на активите на един дял" in c10:
            v = _to_float(c11_raw)
            if v:
                fund.nav_per_unit = v
        elif c10.strip() == "емисионна стойност":
            v = _to_float(c11_raw)
            if v:
                fund.issue_price = v
        elif "до 1 година" in c10:
            v = _to_float(c11_raw)
            if v:
                fund.redemption_price_short = v
        elif "над 1 година" in c10:
            v = _to_float(c11_raw)
            if v:
                fund.redemption_price_long = v
        elif "нетна стойност на активите" in c10 and "един дял" not in c10:
            v = _to_float(c11_raw)
            if v and not fund.nav_total:
                fund.nav_total = v


def parse_fund_file(path: str | Path) -> Fund:
    """Parse a single .xlsx file. The REPORT sheet is the source of truth."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    if "REPORT" not in wb.sheetnames:
        raise ValueError(f"{path.name}: no REPORT sheet")
    fund = _parse_report_sheet(wb["REPORT"], str(path))

    # If FX wasn't captured (some templates), default to 1
    fund.fx_rates.setdefault("USD/EUR", 0.92)
    fund.fx_rates.setdefault("USD/BGN", 0.92 * 1.95583)
    fund.fx_rates.setdefault("EUR/BGN", 1.95583)
    fund.fx_rates.setdefault("BGN/BGN", 1.0)

    return fund


def parse_data_dir(data_dir: str | Path) -> list[Fund]:
    """Parse every .xlsx under data_dir, sorted by reporting date desc, then name."""
    data_dir = Path(data_dir)
    out: list[Fund] = []
    for p in sorted(data_dir.glob("*.xlsx")):
        if p.name.startswith("~$"):  # Excel lock file
            continue
        try:
            out.append(parse_fund_file(p))
        except Exception as exc:  # noqa: BLE001
            print(f"Could not parse {p.name}: {exc}")
    return out


if __name__ == "__main__":
    import sys
    funds = parse_data_dir(sys.argv[1] if len(sys.argv) > 1 else "data")
    for f in funds:
        print(f"{f.fund_name} | {f.reporting_date} | NAV {f.nav_total:,.2f} | "
              f"NAV/u {f.nav_per_unit:.4f} | {len(f.holdings)} holdings")
        by_class: dict[str, float] = {}
        for h in f.holdings:
            by_class[h.asset_class] = by_class.get(h.asset_class, 0) + h.market_value
        for k, v in by_class.items():
            print(f"  {k}: {v:,.2f}")
